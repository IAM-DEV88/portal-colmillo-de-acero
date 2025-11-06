from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Crear archivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Temporada Noviembre 2025"

# Estilos
header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
subheader_fill = PatternFill(start_color="78350f", end_color="78350f", fill_type="solid")
cell_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
font_white = Font(color="FFFFFF", bold=True, size=12)
border = Border(left=Side(style="thin", color="444444"),
                right=Side(style="thin", color="444444"),
                top=Side(style="thin", color="444444"),
                bottom=Side(style="thin", color="444444"))

# Encabezado
ws.merge_cells("A1:D1")
ws["A1"] = "📅 Colmillo de Acero - Temporada Noviembre 2025"
ws["A1"].font = Font(color="FBBF24", bold=True, size=16)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A1"].fill = header_fill

ws.merge_cells("A2:D2")
ws["A2"] = "Horario principal: 18:00 hora servidor | Horario alterno: 00:00 (segundo turno)"
ws["A2"].font = Font(color="FCD34D", italic=True, size=11)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws["A2"].fill = subheader_fill

# Encabezados de tabla
headers = ["Día", "Descripción", "Horario Principal", "Horario Alterno"]
ws.append(headers)
for col in range(1, 5):
    c = ws.cell(row=3, column=col)
    c.font = font_white
    c.fill = subheader_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# Datos
data = [
    ["Miércoles", "Primera ICC 10N semanal (Min. GS 5.5, Máx. 2 cupos bajo GS)", "18:00 ICC 10N", "00:00 Raid Opcional"],
    ["Jueves", "Segunda ICC 10N semanal (Min. GS 5.5, Máx. 2 cupos bajo GS)", "18:00 ICC 10N", "00:00 Raid Opcional"],
    ["Viernes", "Primera ICC 25N semanal (Min. GS 5.7, Máx. 6 cupos bajo GS)", "18:00 ICC 25N", "00:00 Raid Opcional"],
    ["Sábado", "Primera ICC 10H semanal (Min. GS 5.9, Máx. 2 cupos bajo GS)", "18:00 ICC 10H", "00:00 Raid Opcional"],
    ["Domingo", "Segunda ICC 10H semanal (Min. GS 5.9, Máx. 2 cupos bajo GS)", "18:00 ICC 10H", "00:00 Raid Opcional"],
    ["Lunes", "Tercera ICC 10H semanal (Min. GS 5.9, Máx. 2 cupos bajo GS)", "18:00 ICC 10H", "00:00 Raid Opcional"],
    ["Martes", "Primera ICC 25H semanal (Min. GS 6.0, Máx. 4 cupos bajo GS)", "18:00 ICC 25H", "00:00 Raid Opcional"],
]
for row in data:
    ws.append(row)

# Formato
for row in ws.iter_rows(min_row=4, max_row=10, min_col=1, max_col=4):
    for cell in row:
        cell.font = Font(color="E5E7EB", size=11)
        cell.fill = cell_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border

# Tamaño
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22

# Guardar
wb.save("Raids_Temporada_Noviembre_2025.xlsx")
print("Archivo creado: Raids_Temporada_Noviembre_2025.xlsx")
